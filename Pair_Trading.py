"""
Strategy: Statistical Arbitrage via Cointegration-Based Pairs Trading
Spanish Financial Bonds (<6Y Maturity)
Author: Mathieu Dore
Date: November 2025
"""

NOTIONAL_PER_TRADE = 2_000_000
ENTRY_THRESHOLD = 2.0
EXIT_THRESHOLD = 0.5
MAX_NET_DV01 = 2_000

import os
import pandas as pd
import numpy as np
import warnings
from datetime import datetime, timedelta
warnings.filterwarnings('ignore')
nowstr = datetime.now().strftime("%Y%m%d_%H%M%S")

RESULTS_DIR = "results"
os.makedirs(RESULTS_DIR, exist_ok=True)
nowstr = datetime.now().strftime("%Y%m%d_%H%M%S")

import blpapi

from scipy import stats
from statsmodels.tsa.stattools import adfuller, coint
from statsmodels.regression.linear_model import OLS
from statsmodels.tools import add_constant

import matplotlib.pyplot as plt
import seaborn as sns
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("✓ All libraries imported successfully")
print("="*80)

open_trades_df = pd.DataFrame(
    columns=[
        "Pair", "ISIN1", "ISIN2",
        "EntryZScore", "EntryDate", "Direction",
        "NumBonds1", "NumBonds2", "EntryPrice1", "EntryPrice2",
        "CurrentPrice1", "CurrentPrice2",
        "NetDV01", "Notional1", "Notional2",
        "Status", "EntryTime", "LastUpdate", "CurrentZScore"
    ]
)

class BloombergDataFetcher:
    """
    Fetches bond data directly from Bloomberg Terminal via Desktop API
    """
    
    def __init__(self):
        """Initialize Bloomberg session"""
        self.session_options = blpapi.SessionOptions()
        self.session_options.setServerHost("localhost")
        self.session_options.setServerPort(8194)
        self.session = None
        
    def start_session(self):
        """Start Bloomberg session"""
        print("Connecting to Bloomberg Terminal...")
        self.session = blpapi.Session(self.session_options)
        
        if not self.session.start():
            raise Exception("❌ Failed to start Bloomberg session. Ensure Terminal is running and logged in.")
        
        if not self.session.openService("//blp/refdata"):
            self.session.stop()
            raise Exception("❌ Failed to open //blp/refdata service")
        
        print("✓ Connected to Bloomberg Terminal")
        return True
    
    def stop_session(self):
        """Stop Bloomberg session"""
        if self.session:
            self.session.stop()
            print("✓ Bloomberg session closed")
    
    def get_bond_data_for_isins(self, isins, fields=None):

        if fields is None:
            fields = [
                'ID_ISIN',
                'NAME',
                'ISSUER',
                'TICKER',
                'COUNTRY_ISO',
                'CNTRY_OF_RISK',
                'INDUSTRY_SECTOR',
                'MARKET_SECTOR_DES',
                'MATURITY',
                'MTY_TYP',
                'CPN',
                'CPN_FREQ',
                'ISSUE_DT',
                'AMT_OUTSTANDING',
                'COLLAT_TYP',
                'PAYMENT_RANK',
                'SERIES',
                'DUR_MID',
                'DUR_ADJ_MID',
                'CONVEXITY_MID',
                'YLD_YTM_MID',
                'YAS_OAS_SPREAD',
                'Z_SPRD_MID',
                'I_SPRD_MID',
                'ASW_SPREAD',
                'PX_BID',
                'PX_MID',
                'PX_ASK',
                'PX_LAST',
                'RTG_MOODY',
                'RTG_SP',
                'RTG_FITCH',
                'BB_COMPOSITE'
            ]
        
        if not self.session:
            self.start_session()
        
        service = self.session.getService("//blp/refdata")
        
        all_results = []
        batch_size = 50

        for i in range(0, len(isins), batch_size):
            batch = isins[i:i+batch_size]
            
            if i % 100 == 0:
                print(f"Fetching metadata for bonds {i} to {min(i+batch_size, len(isins))}...")
            
            request = service.createRequest("ReferenceDataRequest")
            
            for isin in batch:
                request.append("securities", f"/isin/{isin}")

            for field in fields:
                request.append("fields", field)
            
            self.session.sendRequest(request)

            while True:
                event = self.session.nextEvent()
                
                if event.eventType() == blpapi.Event.RESPONSE or \
                   event.eventType() == blpapi.Event.PARTIAL_RESPONSE:
                    
                    for msg in event:
                        securities = msg.getElement("securityData")
                        
                        for sec in securities.values():
                            security = sec.getElementAsString("security")
                            
                            if sec.hasElement("securityError"):
                                continue
                            
                            field_data = sec.getElement("fieldData")
                            data_dict = {'Security': security}
                            
                            for field in fields:
                                try:
                                    if field_data.hasElement(field):
                                        element = field_data.getElement(field)
                                        if element.isNull():
                                            data_dict[field] = None
                                        else:
                                            data_dict[field] = str(element.getValue())
                                    else:
                                        data_dict[field] = None
                                except:
                                    data_dict[field] = None
                            
                            all_results.append(data_dict)
                
                if event.eventType() == blpapi.Event.RESPONSE:
                    break
        
        df = pd.DataFrame(all_results)
        print(f"✓ Retrieved metadata for {len(df)} bonds")
        
        return df
    
    def get_historical_data(self, isins, start_date, end_date, field='I_SPRD_MID'):
        if not self.session:
            self.start_session()
        
        print(f"\nFetching historical {field} data...")
        print(f"Period: {start_date} to {end_date}")
        print(f"Bonds: {len(isins)}")
        
        all_data = {}
        service = self.session.getService("//blp/refdata")
        
        for idx, isin in enumerate(isins, 1):
            if idx % 20 == 0 or idx == len(isins):
                print(f"Progress: {idx}/{len(isins)} bonds...")
            
            request = service.createRequest("HistoricalDataRequest")
            request.append("securities", f"/isin/{isin}")
            request.append("fields", field)
            request.set("startDate", start_date)
            request.set("endDate", end_date)
            request.set("periodicitySelection", "DAILY")
            
            self.session.sendRequest(request)
            
            time_series = {}
            while True:
                event = self.session.nextEvent()
                
                if event.eventType() == blpapi.Event.RESPONSE or \
                   event.eventType() == blpapi.Event.PARTIAL_RESPONSE:
                    
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        
                        if security_data.hasElement("securityError"):
                            break
                        
                        field_data = security_data.getElement("fieldData")
                        
                        for i in range(field_data.numValues()):
                            field_value = field_data.getValue(i)
                            date = field_value.getElementAsDatetime("date")
                            
                            if field_value.hasElement(field):
                                try:
                                    value = field_value.getElementAsFloat(field)
                                    time_series[pd.to_datetime(str(date))] = value
                                except:
                                    pass
                
                if event.eventType() == blpapi.Event.RESPONSE:
                    break
            
            if time_series:
                all_data[isin] = pd.Series(time_series)
        
        df = pd.DataFrame(all_data)
        df = df.sort_index()
        
        print(f"✓ Retrieved {len(df)} days of data for {len(df.columns)} bonds")
        if len(df) > 0:
            print(f"✓ Date range: {df.index[0].date()} to {df.index[-1].date()}")
            print(f"✓ Missing data: {df.isna().sum().sum()} cells ({df.isna().sum().sum()/df.size*100:.1f}%)")
        
        return df

bb = BloombergDataFetcher()
bb.start_session()
excel_file = "FBES-PAIR-Copy-2.xlsx"
bonds_metadata_excel = pd.read_excel(excel_file, sheet_name="Bonds")
print(bonds_metadata_excel.columns)
bond_isins = bonds_metadata_excel['ISIN'].tolist()

print("First 5 ISINs from Excel:", bond_isins[:5])
print("Total number of ISINs:", len(bond_isins))


# Define date range for historical data
end_date = datetime.now().strftime("%Y%m%d")
start_date = (datetime.now() - timedelta(days=180)).strftime("%Y%m%d")

print(f"\n✓ Fetching Bloomberg data for {{len(bond_isins)}} bonds...")
print(f"  Historical period: {{start_date}} to {{end_date}}")

bondsdf = bb.get_bond_data_for_isins(bond_isins)
ispreadsraw = bb.get_historical_data(bond_isins, start_date, end_date, 'I_SPRD_MID')
pricesraw = bb.get_historical_data(bond_isins, start_date, end_date, 'PX_LAST')
zspreadsraw = bb.get_historical_data(bond_isins, start_date, end_date, 'Z_SPRD_MID')
bb.stop_session()

print("\n✓ All data fetched from Bloomberg Terminal")
print(f"✓ Bonds with metadata: {len(bondsdf)}")
print(f"✓ Historical trading days: {len(ispreadsraw)}")
if len(ispreadsraw) > 0:
    print(f"✓ Date range: {ispreadsraw.index[0].date()} to {ispreadsraw.index[-1].date()}")

class DataProcessor:
    """Clean and prepare Bloomberg API data for analysis"""
    
    @staticmethod
    def clean_historical_data(dfraw, datatype="spreads"):
        df = dfraw.copy()

        missing_pct = df.isna().sum() / len(df)
        valid_cols = missing_pct[missing_pct < 0.5].index
        df = df[valid_cols]

        df = df.ffill(limit=5)
        
        print(f"{datatype.title()}: {df.shape[1]} bonds, {df.shape[0]} days")
        print(f"Missing data: {df.isna().sum().sum()} cells ({df.isna().sum().sum()/df.size*100:.1f}%)")
        
        return df
    
    @staticmethod
    def prepare_bond_metadata(bondsdf):
        """Extract and clean bond static data from Bloomberg API"""
        metadata = bondsdf.copy()
        if 'ID_ISIN' in metadata.columns:
            metadata['ISIN'] = metadata['ID_ISIN']
        else:
            print("Warning: No ISIN field found in metadata")
            return metadata
        
        if 'ISSUER' in metadata.columns:
            metadata['IssuerBase'] = metadata['ISSUER'].str.split().str[0]

        if 'MATURITY' in metadata.columns:
            metadata['MaturityDate'] = pd.to_datetime(metadata['MATURITY'], errors='coerce')
            metadata['YearsToMaturity'] = (metadata['MaturityDate'] - pd.Timestamp.now()).dt.days / 365.25

        numeric_cols = ['CPN', 'DUR_MID', 'YLD_YTM_MID', 'AMT_OUTSTANDING', 
                       'CONVEXITY_MID', 'YAS_OAS_SPREAD', 'Z_SPRD_MID', 'I_SPRD_MID',
                       'PX_BID', 'PX_MID', 'PX_ASK']
        
        for col in numeric_cols:
            if col in metadata.columns:
                metadata[col] = pd.to_numeric(metadata[col], errors='coerce')
        
        # Filter to <6Y maturity
        if 'YearsToMaturity' in metadata.columns:
            metadata = metadata[metadata['YearsToMaturity'] < 6].copy()
        
        print(f"Bond Metadata: {len(metadata)} bonds (<6Y maturity)")
        if 'IssuerBase' in metadata.columns:
            print(f"Unique issuers: {metadata['IssuerBase'].nunique()}")
        if 'PAYMENT_RANK' in metadata.columns:
            print(f"Payment ranks: {metadata['PAYMENT_RANK'].unique()}")
        
        return metadata

print("\n" + "="*80)
print("PROCESSING DATA")
print("="*80)

ispreads = DataProcessor.clean_historical_data(ispreadsraw, "I-Spreads")
prices = DataProcessor.clean_historical_data(pricesraw, "Historical Prices")
zspreads = DataProcessor.clean_historical_data(zspreadsraw, "Z-Spreads")

metadata = DataProcessor.prepare_bond_metadata(bondsdf)
print("Metadata columns:", metadata.columns.tolist())
print("Metadata shape:", metadata.shape)
print(metadata.head())

common_isins = list(set(ispreads.columns) & set(prices.columns) & set(metadata['ID_ISIN']))
print(f"\nFinal universe: {len(common_isins)} bonds with complete data")

ispreads = ispreads[common_isins]
prices = prices[common_isins]
zspreads = zspreads[[col for col in common_isins if col in zspreads.columns]]
metadata = metadata[metadata['ISIN'].isin(common_isins)]

print("✓ Data processing complete")

class PairSelector:
    """Identify and rank potential trading pairs"""

    def __init__(self, spreads_df, metadata_df, min_correlation=0.80):
        self.spreads = spreads_df
        self.metadata = metadata_df
        self.min_correlation = min_correlation
        self.correlation_matrix = None
        self.pairs_ranked = None

    def calculate_correlation_matrix(self):
        """Calculate full correlation matrix"""
        print("\n📊 CALCULATING CORRELATION MATRIX...")
        print("="*80)

        self.correlation_matrix = self.spreads.corr()

        # Summary statistics
        corr_values = self.correlation_matrix.values[np.triu_indices_from(self.correlation_matrix.values, k=1)]

        print(f"   Total possible pairs: {len(corr_values):,}")
        print(f"   Correlation stats:")
        print(f"      Mean: {corr_values.mean():.3f}")
        print(f"      Median: {np.median(corr_values):.3f}")
        print(f"      Std Dev: {corr_values.std():.3f}")
        print(f"      Min: {corr_values.min():.3f}")
        print(f"      Max: {corr_values.max():.3f}")
        print(f"\n   Pairs by correlation threshold:")
        print(f"      >0.90: {(corr_values > 0.90).sum():,}")
        print(f"      >0.80: {(corr_values > 0.80).sum():,}")
        print(f"      >0.70: {(corr_values > 0.70).sum():,}")

        return self.correlation_matrix

    def identify_pairs(self):
        """Extract high-correlation pairs with filters"""
        print(f"\nIDENTIFYING PAIRS (correlation >{self.min_correlation})...")
        print("="*80)

        pairs_list = []
        for i in range(len(self.correlation_matrix.columns)):
            for j in range(i+1, len(self.correlation_matrix.columns)):
                isin1 = self.correlation_matrix.columns[i]
                isin2 = self.correlation_matrix.columns[j]
                corr = self.correlation_matrix.iloc[i, j]

                if corr > self.min_correlation:
                    # Get metadata for both bonds
                    bond1 = self.metadata[self.metadata['ISIN'] == isin1].iloc[0]
                    bond2 = self.metadata[self.metadata['ISIN'] == isin2].iloc[0]

                    # Calculate duration difference
                    dur1 = bond1['DUR_MID']
                    dur2 = bond2['DUR_MID']
                    dur_diff = abs(dur1 - dur2)

                    # Calculate maturity difference
                    mat_diff = abs(bond1['YearsToMaturity'] - bond2['YearsToMaturity'])

                    # Check if same issuer
                    same_issuer = bond1['IssuerBase'] == bond2['IssuerBase']

                    pairs_list.append({
                        'ISIN_1': isin1,
                        'ISIN_2': isin2,
                        'Issuer_1': bond1['ISSUER'],
                        'Issuer_2': bond2['ISSUER'],
                        'Issuer_Base_1': bond1['IssuerBase'],
                        'Issuer_Base_2': bond2['IssuerBase'],
                        'Correlation': corr,
                        'Duration_1': dur1,
                        'Duration_2': dur2,
                        'Duration_Diff': dur_diff,
                        'Maturity_1': bond1['YearsToMaturity'],
                        'Maturity_2': bond2['YearsToMaturity'],
                        'Maturity_Diff': mat_diff,
                        'Same_Issuer': same_issuer,
                        'Seniority_1': bond1['PAYMENT_RANK'],
                        'Seniority_2': bond2['PAYMENT_RANK'],
                        'Same_Seniority': bond1['PAYMENT_RANK'] == bond2['PAYMENT_RANK']
                    })

        self.pairs_ranked = pd.DataFrame(pairs_list)

        self.pairs_ranked['Quality_Score'] = (
            self.pairs_ranked['Correlation'] * 0.40 +
            (1 - self.pairs_ranked['Duration_Diff'] / 5) * 0.30 +
            (1 - self.pairs_ranked['Maturity_Diff'] / 5) * 0.20 +
            self.pairs_ranked['Same_Issuer'].astype(float) * 0.10
        )

        self.pairs_ranked = self.pairs_ranked.sort_values('Quality_Score', ascending=False)
        self.pairs_ranked = self.pairs_ranked.reset_index(drop=True)

        print(f" Identified {len(self.pairs_ranked)} potential pairs")
        print(f"\n   Pair composition:")
        print(f"      Same issuer: {self.pairs_ranked['Same_Issuer'].sum()}")
        print(f"      Cross issuer: {(~self.pairs_ranked['Same_Issuer']).sum()}")
        print(f"      Same seniority: {self.pairs_ranked['Same_Seniority'].sum()}")

        return self.pairs_ranked

    def filter_top_pairs(self, max_duration_diff=0.75, same_seniority=True, top_n=50):
        """Apply additional filters for best pairs"""
        print("\n FILTERING TOP PAIRS...")
        print("="*80)

        filtered = self.pairs_ranked.copy()
        print(filtered.columns)

        filtered = filtered[filtered['Duration_Diff'] <= max_duration_diff]
        if same_seniority:
            filtered = filtered[filtered['Same_Seniority']]
    
        # Exclude same issuer pairs
        filtered = filtered[filtered['Issuer_Base_1'] != filtered['Issuer_Base_2']]

        # Take top N by quality score
        filtered = filtered.head(top_n)

        print(f"Filters applied:")
        print(f"  Max duration diff: {max_duration_diff} years")
        print(f"  Same seniority: {same_seniority}")
        print(f"  Top N: {top_n}")
        print(f"\n✅ Final candidate pairs: {len(filtered)}")

        return filtered

print("\n" + "="*80)
print("START PAIR IDENTIFICATION")
print("="*80)

pair_selector = PairSelector(ispreads, metadata, min_correlation=0.80)
correlation_matrix = pair_selector.calculate_correlation_matrix()
all_pairs = pair_selector.identify_pairs()
top_pairs = pair_selector.filter_top_pairs(max_duration_diff=0.75, top_n=50)

print("\n TOP 10 PAIRS:")
print(top_pairs[['Issuer_Base_1', 'Issuer_Base_2', 'Correlation',
                 'Duration_Diff', 'Maturity_Diff', 'Quality_Score']].head(10).to_string(index=False))


class CointegrationTester:
    """Test pairs for cointegration using Engle-Granger methodology"""

    def __init__(self, spreads_df, prices_df):
        self.spreads = spreads_df
        self.prices = prices_df
        self.results = []

    def test_pair(self, isin1, isin2, method='spreads'):
        if method == 'spreads':
            data = self.spreads[[isin1, isin2]].dropna()
            y = data[isin2].values
            X = data[isin1].values
        else:
            data = self.prices[[isin1, isin2]].dropna()
            y = data[isin2].values
            X = data[isin1].values

        # OLS regression: Bond2 = alpha + beta * Bond1 + residuals
        X_with_const = add_constant(X)
        model = OLS(y, X_with_const).fit()

        residuals = model.resid
        alpha = model.params[0]
        beta = model.params[1]  # This is hedge ratio

        # Test residuals for stationarity
        adf_stat, adf_pvalue, _, _, adf_critical, _ = adfuller(residuals, maxlag=1)

        # Engle-Granger cointegration test
        coint_stat, coint_pvalue, coint_critical = coint(X, y)

        # Spread statistics
        spread_mean = residuals.mean()
        spread_std = residuals.std()
        current_spread = residuals[-1]
        z_score = (current_spread - spread_mean) / spread_std

        # Half-life of mean reversion (Ornstein-Uhlenbeck)
        spread_lag = residuals[:-1]
        spread_diff = np.diff(residuals)
        half_life = -np.log(2) / OLS(spread_diff, add_constant(spread_lag)).fit().params[1]

        return {
            'ISIN_1': isin1,
            'ISIN_2': isin2,
            'Alpha': alpha,
            'Beta': beta,
            'R_Squared': model.rsquared,
            'ADF_Statistic': adf_stat,
            'ADF_PValue': adf_pvalue,
            'Coint_PValue': coint_pvalue,
            'Is_Cointegrated': coint_pvalue < 0.05,
            'Spread_Mean': spread_mean,
            'Spread_Std': spread_std,
            'Current_Spread': current_spread,
            'Current_ZScore': z_score,
            'Half_Life': half_life,
            'Data_Points': len(residuals)
        }

    def test_multiple_pairs(self, pairs_df, max_pairs=50):
        """Test cointegration for multiple pairs"""
        print(f"\nTESTING COINTEGRATION FOR {min(len(pairs_df), max_pairs)} PAIRS...")
        print("="*80)

        results_list = []

        for idx, row in pairs_df.head(max_pairs).iterrows():
            try:
                result = self.test_pair(row['ISIN_1'], row['ISIN_2'])
                results_list.append(result)

                if (idx + 1) % 10 == 0:
                    print(f"   Processed {idx + 1} pairs...")

            except Exception as e:
                print(f"Error testing {row['ISIN_1']} vs {row['ISIN_2']}: {str(e)}")
                continue

        self.results = pd.DataFrame(results_list)

        n_cointegrated = self.results['Is_Cointegrated'].sum()
        print(f"\n   ✅ Cointegration testing complete")
        print(f"      Pairs tested: {len(self.results)}")
        print(f"      Cointegrated pairs (p<0.05): {n_cointegrated} ({n_cointegrated/len(self.results)*100:.1f}%)")
        print(f"      Mean half-life: {self.results['Half_Life'].mean():.1f} days")

        return self.results

print("\n" + "="*80)
print("COINTEGRATION ANALYSIS")
print("="*80)

coint_tester = CointegrationTester(ispreads, prices)
coint_results = coint_tester.test_multiple_pairs(top_pairs, max_pairs=50)

final_pairs = top_pairs.merge(coint_results, on=['ISIN_1', 'ISIN_2'])

tradeable_pairs = final_pairs[final_pairs['Is_Cointegrated']].copy()
tradeable_pairs = tradeable_pairs.sort_values('Quality_Score', ascending=False).reset_index(drop=True)

print(f"\n✅ TRADEABLE PAIRS: {len(tradeable_pairs)}")
print("\n📋 TOP 10 TRADEABLE PAIRS:")
print(tradeable_pairs[['Issuer_Base_1', 'Issuer_Base_2', 'Correlation', 'Beta',
                       'Coint_PValue', 'Current_ZScore', 'Half_Life']].head(10).to_string(index=False))

class PairsTradingStrategy:
    """
    Pairs trading strategy with Z-score based signals

    Rules:
    - Entry: |Z-score| > entry_threshold (default 2.0)
    - Exit: |Z-score| < exit_threshold (default 0.5)
    - Stop Loss: |Z-score| > stop_loss (default 3.5)
    - Duration neutral: positions weighted by duration
    """

    def __init__(self, spreads_df, prices_df, metadata_df,
                 entry_threshold=2.0, exit_threshold=0.5, stop_loss=3.5,
                 transaction_cost_bps=10):
        self.spreads = spreads_df
        self.prices = prices_df
        self.metadata = metadata_df
        self.entry_threshold = entry_threshold
        self.exit_threshold = exit_threshold
        self.stop_loss = stop_loss
        self.transaction_cost = transaction_cost_bps / 10000  # bp to fraction

    def calculate_spread_series(self, isin1, isin2, alpha, beta):
        spread_data = self.spreads[[isin1, isin2]].dropna()
        spread = spread_data[isin2] - (alpha + beta * spread_data[isin1])
        return spread

    def calculate_zscores(self, spread, lookback=60):
        spread_mean = spread.rolling(lookback).mean()
        spread_std = spread.rolling(lookback).std()
        zscores = (spread - spread_mean) / spread_std
        return zscores

    def generate_signals(self, zscores):
        signals = pd.Series(0, index=zscores.index)
        position = 0
        for i in range(len(zscores)):
            z = zscores.iloc[i]
            if pd.isna(z):
                signals.iloc[i] = position
                continue
            if position == 0:
                if z > self.entry_threshold:
                    position = -1
                elif z < -self.entry_threshold:
                    position = 1
            elif position != 0:
                if (position == 1 and z < -self.stop_loss) or \
                   (position == -1 and z > self.stop_loss):
                    position = 0
                elif abs(z) < self.exit_threshold:
                    position = 0
            signals.iloc[i] = position
        return signals

    def dynamic_position_sizing(self, spread_series, max_position=1.0):
        vol = spread_series.rolling(window=20).std()
        current_vol = vol.iloc[-1]
        base_size = max_position
        size = base_size / (current_vol if current_vol > 0 else 1)
        size = min(size, max_position)
        return size

    def backtest_pair(self, pair_info, notional=1000000, lookback=60):
        isin1 = pair_info['ISIN_1']
        isin2 = pair_info['ISIN_2']
        alpha = pair_info['Alpha']
        beta = pair_info['Beta']
        dur1 = pair_info['Duration_1']
        dur2 = pair_info['Duration_2']

        spread = self.calculate_spread_series(isin1, isin2, alpha, beta)
        zscores = self.calculate_zscores(spread, lookback)
        signals = self.generate_signals(zscores)
        price_data = self.prices[[isin1, isin2]].reindex(spread.index)
        price_data = price_data.fillna(method='ffill')
        returns_bond1 = price_data[isin1].pct_change()
        returns_bond2 = price_data[isin2].pct_change()
        duration_hedge = dur1 / dur2
        strategy_returns = pd.Series(0.0, index=returns_bond1.index)
        for i in range(1, len(signals)):
            pos = signals.iloc[i-1]
            if pos == 1:
                strategy_returns.iloc[i] = -returns_bond1.iloc[i] + duration_hedge * returns_bond2.iloc[i]
            elif pos == -1:
                strategy_returns.iloc[i] = returns_bond1.iloc[i] - duration_hedge * returns_bond2.iloc[i]
        position_changes = signals.diff().abs()
        transaction_costs = position_changes * self.transaction_cost
        strategy_returns = strategy_returns - transaction_costs
        cumulative_returns = (1 + strategy_returns).cumprod() - 1
        pnl = cumulative_returns * notional
        results = pd.DataFrame({
            'Date': spread.index,
            'Spread': spread.values,
            'ZScore': zscores.values,
            'Signal': signals.values,
            'Price_1': price_data[isin1].values,
            'Price_2': price_data[isin2].values,
            'Returns_1': returns_bond1.values,
            'Returns_2': returns_bond2.values,
            'Strategy_Returns': strategy_returns.values,
            'Cumulative_Returns': cumulative_returns.values,
            'PnL': pnl.values
        })
        return results

    def calculate_performance_metrics(self, backtest_results):
        returns = backtest_results['Strategy_Returns'].dropna()
        if len(returns) == 0 or returns.std() == 0:
            return {}
        total_return = backtest_results['Cumulative_Returns'].iloc[-1]
        annualized_return = (1 + total_return) ** (252 / len(returns)) - 1
        annualized_vol = returns.std() * np.sqrt(252)
        sharpe = annualized_return / annualized_vol if annualized_vol > 0 else 0
        downside_returns = returns[returns < 0]
        downside_std = downside_returns.std() * np.sqrt(252)
        sortino = annualized_return / downside_std if downside_std > 0 else 0
        cumulative = (1 + returns).cumprod()
        running_max = cumulative.expanding().max()
        drawdown = (cumulative - running_max) / running_max
        max_drawdown = drawdown.min()
        win_rate = (returns > 0).sum() / len(returns)
        gross_profit = returns[returns > 0].sum()
        gross_loss = abs(returns[returns < 0].sum())
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else np.inf
        var_95 = returns.quantile(0.05)
        cvar_95 = returns[returns <= var_95].mean()
        signals = backtest_results['Signal']
        trades = (signals.diff() != 0).sum()
        position_lengths = []
        current_length = 0
        for sig in signals:
            if sig != 0:
                current_length += 1
            elif current_length > 0:
                position_lengths.append(current_length)
                current_length = 0
        avg_holding_period = np.mean(position_lengths) if position_lengths else 0
        return {
            'Total_Return': total_return,
            'Annualized_Return': annualized_return,
            'Annualized_Volatility': annualized_vol,
            'Sharpe_Ratio': sharpe,
            'Sortino_Ratio': sortino,
            'Max_Drawdown': max_drawdown,
            'Win_Rate': win_rate,
            'Profit_Factor': profit_factor,
            'VaR_95': var_95,
            'CVaR_95': cvar_95,
            'Number_of_Trades': trades,
            'Avg_Holding_Period': avg_holding_period
        }

print("\n" + "="*80)
print("BACKTESTING STRATEGY")
print("="*80)

strategy = PairsTradingStrategy(
    ispreads, prices, metadata,
    entry_threshold=2.0,
    exit_threshold=0.5,
    stop_loss=3.5,
    transaction_cost_bps=10  # 10 bps round-trip
)

print(f"\nRunning backtests for {len(tradeable_pairs)} pairs...")
print("="*80)

backtest_results_all = {}
performance_metrics_all = []

for idx, pair in tradeable_pairs.iterrows():
    try:
        bt_result = strategy.backtest_pair(pair, notional=1000000, lookback=60)
        metrics = strategy.calculate_performance_metrics(bt_result)

        if metrics:
            metrics['Pair_ID'] = idx
            metrics['ISIN_1'] = pair['ISIN_1']
            metrics['ISIN_2'] = pair['ISIN_2']
            metrics['Issuer_1'] = pair['Issuer_Base_1']
            metrics['Issuer_2'] = pair['Issuer_Base_2']
            metrics['Correlation'] = pair['Correlation']
            metrics['Beta'] = pair['Beta']

            performance_metrics_all.append(metrics)
            backtest_results_all[idx] = bt_result

        if (idx + 1) % 5 == 0:
            print(f"   Completed {idx + 1}/{len(tradeable_pairs)} pairs")

    except Exception as e:
        print(f" Error backtesting pair {idx}: {str(e)}")
        continue

performance_df = pd.DataFrame(performance_metrics_all)

print(f"\nBacktesting complete: {len(performance_df)} pairs")

# Rank by Sharpe ratio
performance_df = performance_df.sort_values('Sharpe_Ratio', ascending=False).reset_index(drop=True)

print("\n" + "="*80)
print("PERFORMANCE ANALYSIS")
print("="*80)

print("\nTOP 10 PAIRS BY SHARPE RATIO:")
print("="*80)
top_performers = performance_df.head(10)[['Issuer_1', 'Issuer_2', 'Sharpe_Ratio',
                                          'Annualized_Return', 'Max_Drawdown',
                                          'Win_Rate', 'Number_of_Trades']]
print(top_performers.to_string(index=False))

print("\nAGGREGATE PORTFOLIO STATISTICS:")
print("="*80)
print(f"Number of pairs analyzed: {len(performance_df)}")
print(f"\nSharpe Ratio Distribution:")
print(f"  Mean: {performance_df['Sharpe_Ratio'].mean():.2f}")
print(f"  Median: {performance_df['Sharpe_Ratio'].median():.2f}")
print(f"  Positive Sharpe: {(performance_df['Sharpe_Ratio'] > 0).sum()} pairs ({(performance_df['Sharpe_Ratio'] > 0).sum()/len(performance_df)*100:.1f}%)")
print(f"  Sharpe > 1.0: {(performance_df['Sharpe_Ratio'] > 1.0).sum()} pairs ({(performance_df['Sharpe_Ratio'] > 1.0).sum()/len(performance_df)*100:.1f}%)")
print(f"\nReturn Metrics:")
print(f"  Mean Ann. Return: {performance_df['Annualized_Return'].mean()*100:.2f}%")
print(f"  Median Ann. Return: {performance_df['Annualized_Return'].median()*100:.2f}%")
print(f"  Best Ann. Return: {performance_df['Annualized_Return'].max()*100:.2f}%")
print(f"  Worst Ann. Return: {performance_df['Annualized_Return'].min()*100:.2f}%")
print(f"\nRisk Metrics:")
print(f"  Mean Max Drawdown: {performance_df['Max_Drawdown'].mean()*100:.2f}%")
print(f"  Mean Win Rate: {performance_df['Win_Rate'].mean()*100:.1f}%")
print(f"  Mean Profit Factor: {performance_df['Profit_Factor'].mean():.2f}")

fig, axes = plt.subplots(2, 3, figsize=(18, 10))
fig.suptitle('Pairs Trading Strategy Performance Analytics', fontsize=16, fontweight='bold')

axes[0, 0].hist(performance_df['Sharpe_Ratio'], bins=30, edgecolor='black', alpha=0.7)
axes[0, 0].axvline(performance_df['Sharpe_Ratio'].median(), color='red',
                   linestyle='--', label=f'Median: {performance_df["Sharpe_Ratio"].median():.2f}')
axes[0, 0].set_xlabel('Sharpe Ratio')
axes[0, 0].set_ylabel('Frequency')
axes[0, 0].set_title('Sharpe Ratio Distribution')
axes[0, 0].legend()
axes[0, 0].grid(alpha=0.3)

axes[0, 1].scatter(performance_df['Annualized_Volatility']*100,
                   performance_df['Annualized_Return']*100,
                   alpha=0.6, s=50)
axes[0, 1].set_xlabel('Annualized Volatility (%)')
axes[0, 1].set_ylabel('Annualized Return (%)')
axes[0, 1].set_title('Risk-Return Profile')
axes[0, 1].grid(alpha=0.3)
axes[0, 1].axhline(0, color='black', linestyle='-', linewidth=0.5)
axes[0, 1].axvline(0, color='black', linestyle='-', linewidth=0.5)

axes[0, 2].hist(performance_df['Max_Drawdown']*100, bins=30,
                edgecolor='black', alpha=0.7, color='coral')
axes[0, 2].axvline(performance_df['Max_Drawdown'].median()*100,
                   color='red', linestyle='--',
                   label=f'Median: {performance_df["Max_Drawdown"].median()*100:.1f}%')
axes[0, 2].set_xlabel('Maximum Drawdown (%)')
axes[0, 2].set_ylabel('Frequency')
axes[0, 2].set_title('Maximum Drawdown Distribution')
axes[0, 2].legend()
axes[0, 2].grid(alpha=0.3)

axes[1, 0].hist(performance_df['Win_Rate']*100, bins=30,
                edgecolor='black', alpha=0.7, color='lightgreen')
axes[1, 0].axvline(performance_df['Win_Rate'].median()*100,
                   color='red', linestyle='--',
                   label=f'Median: {performance_df["Win_Rate"].median()*100:.1f}%')
axes[1, 0].set_xlabel('Win Rate (%)')
axes[1, 0].set_ylabel('Frequency')
axes[1, 0].set_title('Win Rate Distribution')
axes[1, 0].legend()
axes[1, 0].grid(alpha=0.3)

axes[1, 1].hist(performance_df['Number_of_Trades'], bins=30,
                edgecolor='black', alpha=0.7, color='skyblue')
axes[1, 1].axvline(performance_df['Number_of_Trades'].median(),
                   color='red', linestyle='--',
                   label=f'Median: {performance_df["Number_of_Trades"].median():.0f}')
axes[1, 1].set_xlabel('Number of Trades')
axes[1, 1].set_ylabel('Frequency')
axes[1, 1].set_title('Trading Frequency Distribution')
axes[1, 1].legend()
axes[1, 1].grid(alpha=0.3)

axes[1, 2].scatter(performance_df['Correlation'],
                   performance_df['Sharpe_Ratio'],
                   alpha=0.6, s=50, color='purple')
axes[1, 2].set_xlabel('Correlation')
axes[1, 2].set_ylabel('Sharpe Ratio')
axes[1, 2].set_title('Correlation vs Strategy Performance')
axes[1, 2].grid(alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(RESULTS_DIR, f'performance_analytics_{nowstr}.png'), dpi=300, bbox_inches='tight')
plt.show()

print("\nPerformance analytics chart saved: performance_analytics_{nowstr}.png")

print("\n" + "="*80)
print("DETAILED ANALYSIS OF BEST PERFORMING PAIR")
print("="*80)

if len(performance_df) > 0 and len(backtest_results_all) > 0:
    best_pair_idx = performance_df.iloc[0]['Pair_ID']
    best_pair_info = tradeable_pairs[tradeable_pairs.index == best_pair_idx].iloc[0]
    best_backtest = backtest_results_all[best_pair_idx]

    print(f"\n BEST PAIR:")
    print(f"   Issuer 1: {best_pair_info['Issuer_Base_1']}")
    print(f"   Issuer 2: {best_pair_info['Issuer_Base_2']}")
    print(f"   ISIN 1: {best_pair_info['ISIN_1']}")
    print(f"   ISIN 2: {best_pair_info['ISIN_2']}")
    print(f"   Correlation: {best_pair_info['Correlation']:.3f}")
    print(f"   Hedge Ratio (Beta): {best_pair_info['Beta']:.3f}")
    print(f"   Duration 1: {best_pair_info['Duration_1']:.2f} years")
    print(f"   Duration 2: {best_pair_info['Duration_2']:.2f} years")

    best_metrics = performance_df.iloc[0]
    print(f"\n PERFORMANCE METRICS:")
    print(f"   Sharpe Ratio: {best_metrics['Sharpe_Ratio']:.2f}")
    print(f"   Sortino Ratio: {best_metrics['Sortino_Ratio']:.2f}")
    print(f"   Annualized Return: {best_metrics['Annualized_Return']*100:.2f}%")
    print(f"   Annualized Volatility: {best_metrics['Annualized_Volatility']*100:.2f}%")
    print(f"   Maximum Drawdown: {best_metrics['Max_Drawdown']*100:.2f}%")
    print(f"   Win Rate: {best_metrics['Win_Rate']*100:.1f}%")
    print(f"   Profit Factor: {best_metrics['Profit_Factor']:.2f}")
    print(f"   Number of Trades: {int(best_metrics['Number_of_Trades'])}")
    print(f"   Avg Holding Period: {best_metrics['Avg_Holding_Period']:.1f} days")

    plot_data = best_backtest.dropna()

    if len(plot_data) > 10:
        fig, axes = plt.subplots(4, 1, figsize=(16, 14))
        fig.suptitle(f'Best Pair: {best_pair_info["Issuer_Base_1"]} vs {best_pair_info["Issuer_Base_2"]}',
                     fontsize=16, fontweight='bold')

        ax1 = axes[0]
        ax1_twin = ax1.twinx()
        ax1.plot(plot_data['Date'], plot_data['Spread'], color='blue', linewidth=1.5, label='Spread')
        ax1_twin.plot(plot_data['Date'], plot_data['ZScore'], color='red', linewidth=1.5, label='Z-Score')
        ax1_twin.axhline(2.0, color='green', linestyle='--', alpha=0.5, label='Entry Threshold')
        ax1_twin.axhline(-2.0, color='green', linestyle='--', alpha=0.5)
        ax1_twin.axhline(0.5, color='orange', linestyle='--', alpha=0.5, label='Exit Threshold')
        ax1_twin.axhline(-0.5, color='orange', linestyle='--', alpha=0.5)
        ax1.set_ylabel('Spread (bps)', color='blue')
        ax1_twin.set_ylabel('Z-Score', color='red')
        ax1.tick_params(axis='y', labelcolor='blue')
        ax1_twin.tick_params(axis='y', labelcolor='red')
        ax1.set_title('Spread Evolution and Z-Score')
        ax1.grid(alpha=0.3)
        ax1.legend(loc='upper left')
        ax1_twin.legend(loc='upper right')
        ax2 = axes[1]
        signal_colors = plot_data['Signal'].map({1: 'green', -1: 'red', 0: 'gray'})
        ax2.scatter(plot_data['Date'], plot_data['Signal'], c=signal_colors, alpha=0.6, s=20)
        ax2.set_ylabel('Position')
        ax2.set_yticks([-1, 0, 1])
        ax2.set_yticklabels(['Short Spread', 'Flat', 'Long Spread'])
        ax2.set_title('Trading Signals Over Time')
        ax2.grid(alpha=0.3)
        ax3 = axes[2]
        ax3.plot(plot_data['Date'], plot_data['PnL'], color='darkgreen', linewidth=2)
        ax3.fill_between(plot_data['Date'], 0, plot_data['PnL'],
                         where=(plot_data['PnL'] >= 0), alpha=0.3, color='green', label='Profit')
        ax3.fill_between(plot_data['Date'], 0, plot_data['PnL'],
                         where=(plot_data['PnL'] < 0), alpha=0.3, color='red', label='Loss')
        ax3.axhline(0, color='black', linestyle='-', linewidth=0.5)
        ax3.set_ylabel('P&L (EUR)')
        final_pnl = plot_data['PnL'].iloc[-1]
        ax3.set_title(f'Cumulative P&L (Final: €{final_pnl:,.0f})')
        ax3.legend()
        ax3.grid(alpha=0.3)
        ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x/1000:.0f}K'))

        # Rolling Sharpe Ratio (90-day)
        ax4 = axes[3]
        rolling_returns = plot_data['Strategy_Returns'].rolling(90)
        rolling_sharpe = (rolling_returns.mean() / rolling_returns.std()) * np.sqrt(252)
        ax4.plot(plot_data['Date'], rolling_sharpe, color='purple', linewidth=2)
        ax4.axhline(0, color='black', linestyle='-', linewidth=0.5)
        ax4.axhline(1.0, color='green', linestyle='--', alpha=0.5, label='Sharpe = 1.0')
        ax4.set_ylabel('Rolling Sharpe Ratio')
        ax4.set_xlabel('Date')
        ax4.set_title('90-Day Rolling Sharpe Ratio')
        ax4.legend()
        ax4.grid(alpha=0.3)

        plt.tight_layout()
        plt.savefig(os.path.join(RESULTS_DIR, f'best_pair_analysis_{nowstr}.png'), dpi=300, bbox_inches='tight')
        plt.show()

        print("\n Best pair analysis chart saved: best_pair_analysis_{nowstr}.png")
    else:
        print("\n WARNING: Insufficient data points for detailed visualization")
        print(f"   Available data points: {len(plot_data)}")
else:
    print("\n WARNING: No pairs passed performance criteria for detailed analysis")

print("\n" + "="*80)
print("PORTFOLIO CONSTRUCTION")
print("="*80)

TOP_N_PAIRS = 10
NOTIONAL_PER_PAIR = 1000000  # €1M per pair

portfolio_pairs = performance_df[performance_df['Sharpe_Ratio'] > 1.0].head(TOP_N_PAIRS).copy()

if len(portfolio_pairs) == 0:
    print("\n WARNING: No pairs with Sharpe > 1.0. Using top 5 pairs instead.")
    portfolio_pairs = performance_df.head(5).copy()

print(f"\n PORTFOLIO COMPOSITION: Top {len(portfolio_pairs)} Pairs")
print("="*80)
print(portfolio_pairs[['Issuer_1', 'Issuer_2', 'Sharpe_Ratio', 'Annualized_Return',
                       'Max_Drawdown', 'Number_of_Trades']].to_string(index=False))
portfolio_pnl = None

for idx, pair in portfolio_pairs.iterrows():
    pair_id = pair['Pair_ID']

    if pair_id not in backtest_results_all:
        continue

    bt_result = backtest_results_all[pair_id].copy()
    weight = 1 / len(portfolio_pairs)

    bt_result['Weighted_PnL'] = bt_result['PnL'] * weight
    bt_result['Weighted_Returns'] = bt_result['Strategy_Returns'] * weight

    if portfolio_pnl is None:
        portfolio_pnl = bt_result[['Date', 'Weighted_PnL', 'Weighted_Returns']].copy()
        portfolio_pnl.columns = ['Date', 'Total_PnL', 'Total_Returns']
    else:
        temp_df = bt_result[['Date', 'Weighted_PnL', 'Weighted_Returns']].copy()
        temp_df.columns = ['Date', 'PnL_temp', 'Returns_temp']
        portfolio_pnl = portfolio_pnl.merge(temp_df, on='Date', how='outer')
        portfolio_pnl['Total_PnL'] = portfolio_pnl['Total_PnL'].fillna(0) + portfolio_pnl['PnL_temp'].fillna(0)
        portfolio_pnl['Total_Returns'] = portfolio_pnl['Total_Returns'].fillna(0) + portfolio_pnl['Returns_temp'].fillna(0)
        portfolio_pnl = portfolio_pnl[['Date', 'Total_PnL', 'Total_Returns']]

if portfolio_pnl is not None:
    portfolio_pnl = portfolio_pnl.sort_values('Date').drop_duplicates(subset=['Date']).reset_index(drop=True)

    print(f"\n Portfolio P&L Data Points: {len(portfolio_pnl)}")
    portfolio_returns = portfolio_pnl['Total_Returns'].dropna()

    if len(portfolio_returns) > 0:
        portfolio_total_return = (1 + portfolio_returns).prod() - 1
        portfolio_ann_return = (1 + portfolio_total_return) ** (252 / len(portfolio_returns)) - 1
        portfolio_ann_vol = portfolio_returns.std() * np.sqrt(252)
        portfolio_sharpe = portfolio_ann_return / portfolio_ann_vol if portfolio_ann_vol > 0 else 0
        portfolio_cumret = (1 + portfolio_returns).cumprod()
        portfolio_running_max = portfolio_cumret.expanding().max()
        portfolio_drawdown = (portfolio_cumret - portfolio_running_max) / portfolio_running_max
        portfolio_max_dd = portfolio_drawdown.min()

        print(f"\n PORTFOLIO PERFORMANCE:")
        print("="*80)
        print(f"Total Notional: €{NOTIONAL_PER_PAIR * len(portfolio_pairs):,.0f}")
        print(f"Number of Pairs: {len(portfolio_pairs)}")
        print(f"Total Return: {portfolio_total_return*100:.2f}%")
        print(f"Annualized Return: {portfolio_ann_return*100:.2f}%")
        print(f"Annualized Volatility: {portfolio_ann_vol*100:.2f}%")
        print(f"Sharpe Ratio: {portfolio_sharpe:.2f}")
        print(f"Maximum Drawdown: {portfolio_max_dd*100:.2f}%")
        print(f"Final P&L: €{portfolio_pnl['Total_PnL'].iloc[-1]:,.0f}")

        fig, axes = plt.subplots(2, 1, figsize=(16, 10))
        fig.suptitle(f'Portfolio Performance ({len(portfolio_pairs)} Pairs)',
                     fontsize=16, fontweight='bold')
        axes[0].plot(portfolio_pnl['Date'], portfolio_pnl['Total_PnL'],
                     color='darkblue', linewidth=2.5)
        axes[0].fill_between(portfolio_pnl['Date'], 0, portfolio_pnl['Total_PnL'],
                              where=(portfolio_pnl['Total_PnL'] >= 0),
                              alpha=0.3, color='green')
        axes[0].fill_between(portfolio_pnl['Date'], 0, portfolio_pnl['Total_PnL'],
                              where=(portfolio_pnl['Total_PnL'] < 0),
                              alpha=0.3, color='red')
        axes[0].axhline(0, color='black', linestyle='-', linewidth=0.5)
        axes[0].set_ylabel('Portfolio P&L (EUR)')
        axes[0].set_title(f'Cumulative Portfolio P&L (Final: €{portfolio_pnl["Total_PnL"].iloc[-1]:,.0f})')
        axes[0].grid(alpha=0.3)
        axes[0].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x/1000:.0f}K'))

        axes[1].fill_between(range(len(portfolio_drawdown)), 0, portfolio_drawdown.values*100,
                              alpha=0.5, color='red')
        axes[1].plot(portfolio_drawdown.values*100, color='darkred', linewidth=2)
        axes[1].set_ylabel('Drawdown (%)')
        axes[1].set_xlabel('Trading Days')
        axes[1].set_title(f'Portfolio Drawdown (Max: {portfolio_max_dd*100:.2f}%)')
        axes[1].grid(alpha=0.3)

        plt.tight_layout()
        plt.savefig(os.path.join(RESULTS_DIR, f'portfolio_performance_{nowstr}.png'), dpi=300, bbox_inches='tight')
        plt.show()

        print("\n Portfolio performance chart saved: portfolio_performance_{nowstr}.png")
    else:
        print("\n No valid portfolio returns to analyze")
        portfolio_sharpe = 0
        portfolio_ann_return = 0
        portfolio_max_dd = 0
else:
    print("\n Could not construct portfolio P&L")
    portfolio_pnl = pd.DataFrame(columns=['Date', 'Total_PnL', 'Total_Returns'])

print("\n" + "="*80)
print("CURRENT TRADING SIGNALS")
print("="*80)

current_signals = []

for idx, pair in portfolio_pairs.iterrows():
    pair_id = pair['Pair_ID']
    isin1 = pair['ISIN_1']
    isin2 = pair['ISIN_2']
    pair_full_info = tradeable_pairs[tradeable_pairs.index == pair_id].iloc[0]
    alpha = pair_full_info['Alpha']
    beta = pair_full_info['Beta']

    try:
        spread = strategy.calculate_spread_series(isin1, isin2, alpha, beta)
        zscores = strategy.calculate_zscores(spread, lookback=60)
        current_z = zscores.iloc[-1] if len(zscores) > 0 and not pd.isna(zscores.iloc[-1]) else 0
        current_spread = spread.iloc[-1] if len(spread) > 0 else 0

        if current_z > strategy.entry_threshold:
            signal = "SHORT SPREAD"
            action = f"Short {pair['Issuer_1']}, Long {pair['Issuer_2']}"
            confidence = "HIGH" if abs(current_z) > 2.5 else "MEDIUM"
        elif current_z < -strategy.entry_threshold:
            signal = "LONG SPREAD"
            action = f"Long {pair['Issuer_1']}, Short {pair['Issuer_2']}"
            confidence = "HIGH" if abs(current_z) > 2.5 else "MEDIUM"
        elif abs(current_z) < strategy.exit_threshold:
            signal = "FLAT"
            action = "Close existing positions"
            confidence = "N/A"
        else:
            signal = "HOLD"
            action = "Maintain current position"
            confidence = "LOW"

        current_signals.append({
            'Pair': f"{pair['Issuer_1']} / {pair['Issuer_2']}",
            'ISIN_1': isin1,
            'ISIN_2': isin2,
            'Current_ZScore': current_z,
            'Signal': signal,
            'Action': action,
            'Confidence': confidence,
            'Beta': beta,
            'Sharpe_Ratio': pair['Sharpe_Ratio']
        })

    except Exception as e:
        print(f" Error calculating signal for pair {idx}: {str(e)}")
        continue

if len(current_signals) > 0:
    signals_df = pd.DataFrame(current_signals)

    print(f"\n LIVE TRADING SIGNALS (as of latest data):")
    print("="*80)
    print(signals_df[['Pair', 'Current_ZScore', 'Signal', 'Confidence']].to_string(index=False))
    active_signals = signals_df[signals_df['Signal'].isin(['LONG SPREAD', 'SHORT SPREAD'])]

def print_trade_suggestions(signals_df, metadata):
    """Enhanced trading desk signal output"""
    print("\n ACTIONABLE TRADE SIGNALS (ENTRY OPPORTUNITIES)")
    print("="*80)

    actionable = signals_df[signals_df['Signal'].isin(['LONG SPREAD', 'SHORT SPREAD'])]
    if len(actionable) == 0:
        print(" No new entry signals at current levels.")
        print("  → Monitor pairs below for upcoming opportunities\n")
        return

    for idx, sig in actionable.iterrows():
        action_text = "SHORT the spread" if sig['Signal'] == "SHORT SPREAD" else "LONG the spread"
        print(f"\n PAIR #{idx+1}: {sig['Pair']}")
        print(f"   Current Z-Score: {sig['Current_ZScore']:.2f}")
        print(f"   → ACTION: {action_text}")
        print(f"   → Confidence: {sig['Confidence']}")
        print(f"   → Expected Sharpe: {sig['Sharpe_Ratio']:.2f}")
        print(f"   → Hedge Ratio (Beta): {sig['Beta']:.3f}")
        entry_threshold = 2.0
        distance_to_reversal = abs(sig['Current_ZScore']) - entry_threshold
        print(f"   → Signal Strength: {distance_to_reversal:.2f} Z above threshold")

def print_trade_suggestions(signals_df, metadata, notional=1_000_000):
    """Enhanced trading desk signal output incl. position sizing"""
    print("\n ACTIONABLE TRADE SIGNALS (ENTRY OPPORTUNITIES)")
    print("="*80)

    actionable = signals_df[signals_df['Signal'].isin(['LONG SPREAD', 'SHORT SPREAD'])]
    if len(actionable) == 0:
        print(" No new entry signals at current levels.")
        print("  → Monitor pairs below for upcoming opportunities\n")
        return

    for idx, sig in actionable.iterrows():
        action_text = "SHORT the spread" if sig['Signal'] == "SHORT SPREAD" else "LONG the spread"
        print(f"\n PAIR #{idx+1}: {sig['Pair']}")
        print(f"   Current Z-Score: {sig['Current_ZScore']:.2f}")
        print(f"   → ACTION: {action_text}")
        print(f"   → Confidence: {sig['Confidence']}")
        print(f"   → Expected Sharpe: {sig['Sharpe_Ratio']:.2f}")
        print(f"   → Hedge Ratio (Beta): {sig['Beta']:.3f}")

        # Position sizing
        try:
            pos_info = calculate_position_sizes(sig, metadata, notional)
            if pos_info:
                print(f" SUGGESTED SIZING for €{notional:,.0f} notional:")
                print(f"      Bond 1 ({pos_info['ISIN1']}): {pos_info['NumBonds1']:.0f} units @ €{pos_info['Price1']:.2f} | Duration: {pos_info['Duration1']:.2f} | DV01: {pos_info['DV01_1']:.2f}")
                print(f"      Bond 2 ({pos_info['ISIN2']}): {pos_info['NumBonds2']:.0f} units @ €{pos_info['Price2']:.2f} | Duration: {pos_info['Duration2']:.2f} | DV01: {pos_info['DV01_2']:.2f}")
                print(f"      Duration hedge (Net DV01): {pos_info['NetDV01']:.2f}")
        except Exception as e:
            print(f"   [Position sizing unavailable: {e}]")
   
    print("\n" + "="*80)

def print_monitoring_pairs(signals_df):
    print("\n MONITORING LIST (Watch for entry)")
    print("="*80)
    monitoring = signals_df[signals_df['Signal'] == "HOLD"]
    if len(monitoring) == 0:
        print("No pairs in monitoring mode.\n")
        return
    print(f"Tracking {len(monitoring)} pairs for potential entry:\n")
    for idx, sig in monitoring.head(10).iterrows():
        z_val = sig['Current_ZScore']
        print(f"  • {sig['Pair']}: Z={z_val:.2f} (waiting for ±2.0)")
    print("\n" + "="*80)

if not signals_df.empty:
    print_trade_suggestions(signals_df, metadata)
    print_monitoring_pairs(signals_df)
else:
    print("\n No signals could be calculated")
    signals_df = pd.DataFrame()
    active_signals = pd.DataFrame()

print("\n" + "="*80)
print("POSITION SIZING CALCULATOR")
print("="*80)

def calculate_position_sizes(pair_info, metadata, notional=1000000):
    """
    Calculate duration-neutral position sizes

    Parameters:
    -----------
    pair_info : dict/Series
        Pair information
    notional : float
        Target notional for leg 1

    Returns:
    --------
    dict with position details
    """
    isin1 = pair_info['ISIN_1']
    isin2 = pair_info['ISIN_2']
    bond1_matches = metadata[metadata['ISIN'] == isin1]
    bond2_matches = metadata[metadata['ISIN'] == isin2]

    if len(bond1_matches) == 0 or len(bond2_matches) == 0:
        return None

    bond1 = bond1_matches.iloc[0]
    bond2 = bond2_matches.iloc[0]

    dur1 = bond1['Mod Dur (Mid)']
    dur2 = bond2['Mod Dur (Mid)']
    price1 = bond1['Mid Price']
    price2 = bond2['Mid Price']
    beta = pair_info['Beta']

    notional1 = notional
    notional2 = (dur1 * price1) / (dur2 * price2) * notional1

    num_bonds1 = notional1 / (price1 * 10)
    num_bonds2 = notional2 / (price2 * 10)

    dv01_1 = dur1 * price1 / 10000
    dv01_2 = dur2 * price2 / 10000

    return {
        'ISIN_1': isin1,
        'ISIN_2': isin2,
        'Bond_1': bond1['Issuer Name'],
        'Bond_2': bond2['Issuer Name'],
        'Price_1': price1,
        'Price_2': price2,
        'Duration_1': dur1,
        'Duration_2': dur2,
        'Notional_1': notional1,
        'Notional_2': notional2,
        'Num_Bonds_1': num_bonds1,
        'Num_Bonds_2': num_bonds2,
        'DV01_1': dv01_1,
        'DV01_2': dv01_2,
        'Net_DV01': abs(dv01_1 * num_bonds1 - dv01_2 * num_bonds2)
    }

def log_trade_signal(sig, pos_info, filename=os.path.join(RESULTS_DIR, f"TradeBlotter_{nowstr}.csv")):
    import os, csv
    fields = ['datetime', 'pair', 'action', 'notional', 'entry_zscore', 'confidence', 'num_bonds_1', 'num_bonds_2', 'isin1', 'isin2']
    data = [pd.Timestamp.now(), sig['Pair'], sig['Signal'], pos_info['Notional1'], sig['Current_ZScore'], sig['Confidence'], pos_info['NumBonds1'], pos_info['NumBonds2'], pos_info['ISIN1'], pos_info['ISIN2']]
    write_header = not os.path.isfile(filename)
    with open(filename, 'a', newline='') as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(fields)
        writer.writerow(data)

def print_trade_suggestions(signals_df, metadata, notional=1_000_000):
    print("\n ACTIONABLE TRADE SIGNALS (READY TO TRADE)")
    print("="*80)
    actionable = signals_df[signals_df['Signal'].isin(['LONG SPREAD', 'SHORT SPREAD'])]
    if len(actionable) == 0:
        print(" No entry signals at this moment.")
        print("Monitor monitoring list shown below for near-threshold pairs.")
        return
    for idx, sig in actionable.iterrows():
        action_text = "SHORT the spread" if sig['Signal'] == "SHORT SPREAD" else "LONG the spread"
        print(f"\n PAIR #{idx+1}: {sig['Pair']}")
        print(f"   Z-Score: {sig['Current_ZScore']:.2f} | Action: {action_text} | Confidence: {sig['Confidence']}")
        print(f"   Expected Sharpe: {sig['Sharpe_Ratio']:.2f} | Beta (hedge ratio): {sig['Beta']:.3f}")
        try:
            pos_info = calculate_position_sizes(sig, metadata, notional)
            if pos_info:
                print(f" EXECUTION SIZING (€{notional:,.0f}):")
                print(f"      Bond 1: {pos_info['ISIN1']} ({pos_info.get('Bond1','')}) – {pos_info['NumBonds1']:.0f} units @ €{pos_info['Price1']:.2f} | Dur: {pos_info['Duration1']:.2f} | DV01: {pos_info['DV01_1']:.2f}")
                print(f"      Bond 2: {pos_info['ISIN2']} ({pos_info.get('Bond2','')}) – {pos_info['NumBonds2']:.0f} units @ €{pos_info['Price2']:.2f} | Dur: {pos_info['Duration2']:.2f} | DV01: {pos_info['DV01_2']:.2f}")
                print(f"      Net DV01 hedge: {pos_info['NetDV01']:.2f} (lower=better)")
                log_trade_signal(sig, pos_info)
        except Exception as e:
            print(f"   [Position sizing unavailable: {e}]")
    print("="*80)

def print_monitoring_pairs(signals_df):
    print("\n MONITORING LIST (Watch for entry)")
    print("="*80)
    monitoring = signals_df[signals_df['Signal'] == "HOLD"]
    if len(monitoring) == 0:
        print("No pairs in monitoring mode.\n")
        return
    for idx, sig in monitoring.iterrows():
        wait_z = 2 - abs(sig['Current_ZScore'])
        print(f"  • {sig['Pair']}: Z={sig['Current_ZScore']:.2f} (needs {wait_z:+.2f} to trigger)")
    print("="*80)

def print_risk_summary(open_trades_df):
    print("\n RISK SUMMARY:")
    if open_trades_df.empty:
        print("No open trades. Portfolio risk is flat.\n")
        return
    net_dv01 = open_trades_df['NetDV01'].sum()
    top_exposure = open_trades_df['Notional1'].abs().nlargest(3).sum()
    print(f"  Net portfolio DV01: {net_dv01:.1f}")
    print(f"  Largest single trade: €{top_exposure:,.0f}")
    print(f"  Open trades: {len(open_trades_df)}\n")

print("\n" + "="*80)
print("EXPORTING ENHANCED RESULTS")
print("="*80)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    output_file = os.path.join(RESULTS_DIR, f'Pairs_Trading_Results_Enhanced_{nowstr}.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        glossary_data = {
            'Metric': [
                'PERFORMANCE METRICS:',
                'Total Return (%)',
                'Annual Return (%)',
                'Annual Volatility (%)',
                'Sharpe Ratio',
                'Sortino Ratio',
                'Max Drawdown (%)',
                'Win Rate (%)',
                'Profit Factor',
                '',
                'RISK METRICS:',
                'VaR 95% (daily %)',
                'CVaR 95% (daily %)',
                '',
                'PAIR CHARACTERISTICS:',
                'Spread Correlation',
                'Hedge Ratio (β)',
                'Number of Trades',
                'Avg Hold Period (days)',
                '',
                'TRADING SIGNALS:',
                'Z-Score',
                'Signal',
                'Confidence',
                '',
                'POSITION SIZING:',
                'Duration (years)',
                'DV01 (€)',
                'Notional (€)',
                '',
                'STRATEGY PARAMETERS:',
                'Entry Threshold',
                'Exit Threshold',
                'Stop Loss',
                'Transaction Cost',
            ],
            'Definition': [
                '',
                'Cumulative return over entire backtest period',
                'Total return annualized to 252 trading days per year',
                'Standard deviation of daily returns (annualized to 252 days)',
                'Risk-adjusted return: Annual Return / Annual Volatility',
                'Downside risk-adjusted return: only penalizes downside volatility',
                'Largest peak-to-trough decline during backtest (always negative)',
                'Percentage of profitable trades out of total trades',
                'Gross profit divided by gross loss',
                '',
                '',
                'Value at Risk: expected maximum daily loss at 95% confidence level',
                'Conditional VaR (Expected Shortfall): average loss when VaR is exceeded',
                '',
                '',
                'Pearson correlation between the two bond spreads (0 to 1)',
                'OLS regression coefficient: units of Bond 2 needed to hedge 1 unit of Bond 1',
                'Total number of round-trip trades executed during backtest period',
                'Average number of calendar days a position was held before closing',
                '',
                '',
                'Standardized spread deviation from rolling mean',
                'Trading action: LONG SPREAD (buy undervalued) or SHORT SPREAD (sell overvalued)',
                'Signal strength based on Z-score magnitude',
                '',
                '',
                'Modified duration: % price change per 1% yield change',
                'Dollar value of 1 basis point (0.01%) yield change',
                'Total position size in euros for duration-neutral hedging',
                '',
                '',
                'Z-score threshold to enter new positions',
                'Z-score threshold to exit existing positions',
                'Maximum Z-score before forced exit (risk limit)',
                'Estimated round-trip transaction cost in basis points',
            ],
            'Good Value / Target': [
                '',
                '> 10% for 6-month backtest',
                '> 8% annualized',
                '< 15% (lower is better)',
                '> 1.5 (good), > 2.0 (excellent)',
                '> 1.0',
                '> -15% (less negative is better)',
                '> 55%',
                '> 2.0',
                '',
                '',
                'Closer to 0% (less negative)',
                'Closer to 0% (less negative)',
                '',
                '',
                '> 0.85 (highly correlated)',
                'Stable over time (1.0-2.0 typical)',
                '10-30 trades (sufficient sample)',
                '10-40 days (too short = noise, too long = capital inefficiency)',
                '',
                '',
                '|Z| > 2.0 for entry, |Z| < 0.5 for exit',
                'Follow system rules consistently',
                'HIGH (|Z|>2.5) or MEDIUM (2.0<|Z|<2.5) preferred',
                '',
                '',
                'Similar for both bonds (< 0.5 year difference)',
                'Net DV01 ≈ €0 for perfect hedge',
                '€1,000,000 per pair recommended',
                '',
                '',
                '2.0 standard deviations',
                '0.5 standard deviations',
                '3.5 standard deviations',
                '10 bps for Spanish corporate bonds',
            ]
        }

        glossary_df = pd.DataFrame(glossary_data)
        glossary_df.to_excel(writer, sheet_name='README_Glossary', index=False)

        if len(portfolio_pairs) > 0:
            portfolio_enhanced = portfolio_pairs.copy()

            # Rename columns with units
            column_mapping = {
                'Total_Return': 'Total Return (%)',
                'Annualized_Return': 'Annual Return (%)',
                'Annualized_Volatility': 'Annual Volatility (%)',
                'Sharpe_Ratio': 'Sharpe Ratio',
                'Sortino_Ratio': 'Sortino Ratio',
                'Max_Drawdown': 'Max Drawdown (%)',
                'Win_Rate': 'Win Rate (%)',
                'Profit_Factor': 'Profit Factor',
                'VaR_95': 'VaR 95% (daily %)',
                'CVaR_95': 'CVaR 95% (daily %)',
                'Number_of_Trades': 'Number of Trades',
                'Avg_Holding_Period': 'Avg Hold Period (days)',
                'Issuer_1': 'Bond 1 Issuer',
                'Issuer_2': 'Bond 2 Issuer',
                'Correlation': 'Spread Correlation',
                'Beta': 'Hedge Ratio (β)'
            }

            portfolio_enhanced = portfolio_enhanced.rename(columns=column_mapping)

            pct_cols = ['Total Return (%)', 'Annual Return (%)', 'Annual Volatility (%)',
                        'Max Drawdown (%)', 'Win Rate (%)', 'VaR 95% (daily %)', 'CVaR 95% (daily %)']
            for col in pct_cols:
                if col in portfolio_enhanced.columns:
                    portfolio_enhanced[col] = portfolio_enhanced[col] * 100

            for col in portfolio_enhanced.select_dtypes(include=[np.number]).columns:
                if '(%)' in col:
                    portfolio_enhanced[col] = portfolio_enhanced[col].round(2)
                elif 'Ratio' in col or 'β' in col or 'Correlation' in col or 'Factor' in col:
                    portfolio_enhanced[col] = portfolio_enhanced[col].round(3)
                elif 'days' in col or 'Trades' in col:
                    portfolio_enhanced[col] = portfolio_enhanced[col].round(0).astype(int)

            portfolio_enhanced.to_excel(writer, sheet_name='Portfolio_Pairs', index=False)

        if len(performance_df) > 0:
            all_perf_enhanced = performance_df.copy()

            all_perf_enhanced = all_perf_enhanced.rename(columns=column_mapping)

            for col in pct_cols:
                if col in all_perf_enhanced.columns:
                    all_perf_enhanced[col] = all_perf_enhanced[col] * 100

            for col in all_perf_enhanced.select_dtypes(include=[np.number]).columns:
                if '(%)' in col:
                    all_perf_enhanced[col] = all_perf_enhanced[col].round(2)
                elif 'Ratio' in col or 'β' in col or 'Correlation' in col or 'Factor' in col:
                    all_perf_enhanced[col] = all_perf_enhanced[col].round(3)
                elif 'days' in col or 'Trades' in col:
                    all_perf_enhanced[col] = all_perf_enhanced[col].round(1)

            all_perf_enhanced.to_excel(writer, sheet_name='All_Performance', index=False)

        if len(signals_df) > 0:
            signals_enhanced = signals_df.copy()
            signals_enhanced = signals_enhanced.rename(columns={
                'Current_ZScore': 'Current Z-Score',
                'Beta': 'Hedge Ratio (β)',
                'Sharpe_Ratio': 'Expected Sharpe'
            })

            if 'Current Z-Score' in signals_enhanced.columns:
                signals_enhanced['Current Z-Score'] = signals_enhanced['Current Z-Score'].round(2)
            if 'Hedge Ratio (β)' in signals_enhanced.columns:
                signals_enhanced['Hedge Ratio (β)'] = signals_enhanced['Hedge Ratio (β)'].round(3)
            if 'Expected Sharpe' in signals_enhanced.columns:
                signals_enhanced['Expected Sharpe'] = signals_enhanced['Expected Sharpe'].round(2)

            signals_enhanced.to_excel(writer, sheet_name='Current_Signals', index=False)

        if len(backtest_results_all) > 0:
            best_pair_idx = performance_df.iloc[0]['Pair_ID']
            best_backtest = backtest_results_all[best_pair_idx].head(10000).copy()

            best_backtest = best_backtest.rename(columns={
                'ZScore': 'Z-Score',
                'Price_1': 'Bond 1 Price (€)',
                'Price_2': 'Bond 2 Price (€)',
                'Returns_1': 'Bond 1 Daily Return (%)',
                'Returns_2': 'Bond 2 Daily Return (%)',
                'Strategy_Returns': 'Strategy Daily Return (%)',
                'Cumulative_Returns': 'Cumulative Return (%)',
                'PnL': 'P&L (€)'
            })

            ret_cols = ['Bond 1 Daily Return (%)', 'Bond 2 Daily Return (%)',
                        'Strategy Daily Return (%)', 'Cumulative Return (%)']
            for col in ret_cols:
                if col in best_backtest.columns:
                    best_backtest[col] = (best_backtest[col] * 100).round(4)

            if 'Z-Score' in best_backtest.columns:
                best_backtest['Z-Score'] = best_backtest['Z-Score'].round(2)
            if 'Spread' in best_backtest.columns:
                best_backtest['Spread'] = best_backtest['Spread'].round(2)
            if 'P&L (€)' in best_backtest.columns:
                best_backtest['P&L (€)'] = best_backtest['P&L (€)'].round(2)

            best_backtest.to_excel(writer, sheet_name='Best_Pair_Detail', index=False)

        if len(tradeable_pairs) > 0:
            top_50_isins = list(set(
                tradeable_pairs.head(25)['ISIN_1'].tolist() +
                tradeable_pairs.head(25)['ISIN_2'].tolist()
            ))[:50]

            if len(top_50_isins) > 0:
                corr_subset = correlation_matrix.loc[top_50_isins, top_50_isins].round(3)
                corr_subset.to_excel(writer, sheet_name='Correlation_Matrix')

        if portfolio_pnl is not None and len(portfolio_pnl) > 0:
            pnl_enhanced = portfolio_pnl.copy()
            pnl_enhanced = pnl_enhanced.rename(columns={
                'Total_PnL': 'Portfolio P&L (€)',
                'Total_Returns': 'Portfolio Daily Return (%)'
            })

            if 'Portfolio Daily Return (%)' in pnl_enhanced.columns:
                pnl_enhanced['Portfolio Daily Return (%)'] = (pnl_enhanced['Portfolio Daily Return (%)'] * 100).round(4)
            if 'Portfolio P&L (€)' in pnl_enhanced.columns:
                pnl_enhanced['Portfolio P&L (€)'] = pnl_enhanced['Portfolio P&L (€)'].round(2)

            if len(pnl_enhanced) > 100000:
                print(f"  Limiting Portfolio P&L to 100,000 rows (original: {len(pnl_enhanced):,})")
                pnl_enhanced = pnl_enhanced.head(100000)

            pnl_enhanced.to_excel(writer, sheet_name='Portfolio_PnL', index=False)

        strategy_overview = {
            'Parameter': [
                'STRATEGY CONFIGURATION',
                'Entry Threshold',
                'Exit Threshold',
                'Stop Loss Threshold',
                'Transaction Cost',
                'Z-Score Lookback Period',
                'Notional per Pair',
                'Max Pairs in Portfolio',
                '',
                'PAIR SELECTION CRITERIA',
                'Minimum Correlation',
                'Maximum Duration Difference',
                'Cointegration P-Value Max',
                'Same Seniority Required',
                '',
                'STRATEGY LOGIC',
                'Long Spread Entry',
                'Short Spread Entry',
                'Exit Signal',
                'Position Sizing Method',
                '',
                'RISK MANAGEMENT',
                'Duration Neutrality',
                'DV01 Matching Formula',
                'Diversification',
            ],
            'Value': [
                '',
                '2.0 std dev',
                '0.5 std dev',
                '3.5 std dev',
                '10 bps (0.10%)',
                '60 trading days',
                '€1,000,000',
                f'{TOP_N_PAIRS} pairs',
                '',
                '',
                '> 0.80',
                '< 0.75 years',
                '< 0.05 (5%)',
                'Yes',
                '',
                '',
                'Z-Score < -2.0 → Long Bond1, Short Bond2',
                'Z-Score > +2.0 → Short Bond1, Long Bond2',
                '|Z-Score| < 0.5 → Close position',
                'Duration-neutral (Net DV01 ≈ €0)',
                '',
                '',
                'YES - positions weighted by modified duration',
                'Notional₂ = Notional₁ × (Dur₁/Dur₂) × (Price₁/Price₂)',
                'Equal weight across pairs, max 10 pairs',
            ],
            'Explanation': [
                '',
                'Spread must be 2 standard deviations from mean before entering',
                'Close when spread returns within 0.5 std dev of mean (profit-taking)',
                'Force close if spread moves >3.5 std dev against us (loss limit)',
                'Estimated bid-ask spread for Spanish corporate bonds',
                'Calculate rolling mean and std dev over 60 days (~3 months)',
                'Target notional size for each leg of the pair trade',
                'Diversification limit to avoid concentration risk',
                '',
                '',
                'Bond spreads must be highly correlated for stable relationship',
                'Bonds must have similar interest rate sensitivity',
                'Engle-Granger test: pairs must be statistically cointegrated',
                'Only pair bonds with same payment rank (Sr Preferred, Sub, etc.)',
                '',
                '',
                'Spread is cheap vs history - expect it to widen (mean reversion)',
                'Spread is expensive vs history - expect it to narrow (mean reversion)',
                'Spread has mean-reverted back to equilibrium - take profit',
                'Match dollar-duration exposure so rate moves cancel out',
                '',
                '',
                'Ensures position is neutral to parallel yield curve shifts',
                'Precise formula to match interest rate risk between bonds',
                'Spread exposure across multiple pairs reduces idiosyncratic risk',
            ]
        }

        strategy_df = pd.DataFrame(strategy_overview)
        strategy_df.to_excel(writer, sheet_name='Strategy_Overview', index=False)

        summary_data = {
            'Metric': [
                'Portfolio Composition',
                'Number of Pairs',
                'Total Notional',
                'Notional per Pair',
                '',
                'Performance Summary',
                'Total Portfolio Return',
                'Annualized Return',
                'Annualized Volatility',
                'Sharpe Ratio',
                'Maximum Drawdown',
                '',
                'Universe Statistics',
                'Total Bonds Analyzed',
                'High Correlation Pairs (>0.80)',
                'Cointegrated Pairs',
                'Pairs with Sharpe > 1.0',
                '',
                'Current Status',
                'Active Trading Signals',
                'Long Spread Signals',
                'Short Spread Signals',
            ],
            'Value': [
                '',
                f'{len(portfolio_pairs)}',
                f'€{NOTIONAL_PER_PAIR * len(portfolio_pairs):,.0f}',
                f'€{NOTIONAL_PER_PAIR:,.0f}',
                '',
                '',
                f'{portfolio_total_return*100:.2f}%' if 'portfolio_total_return' in locals() else 'N/A',
                f'{portfolio_ann_return*100:.2f}%' if 'portfolio_ann_return' in locals() else 'N/A',
                f'{portfolio_ann_vol*100:.2f}%' if 'portfolio_ann_vol' in locals() else 'N/A',
                f'{portfolio_sharpe:.2f}' if 'portfolio_sharpe' in locals() else 'N/A',
                f'{portfolio_max_dd*100:.2f}%' if 'portfolio_max_dd' in locals() else 'N/A',
                '',
                '',
                f'{len(common_isins)}',
                f'{len(all_pairs)}',
                f'{len(tradeable_pairs)}',
                f'{(performance_df["Sharpe_Ratio"] > 1.0).sum()}',
                '',
                '',
                f'{len(active_signals) if len(signals_df) > 0 else 0}',
                f'{len(signals_df[signals_df["Signal"] == "LONG SPREAD"]) if len(signals_df) > 0 else 0}',
                f'{len(signals_df[signals_df["Signal"] == "SHORT SPREAD"]) if len(signals_df) > 0 else 0}',
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Portfolio_Summary', index=False)

    print(f"\n Enhanced Excel file created: {output_file}")

    wb = load_workbook(output_file)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    title_font = Font(bold=True, size=12)
    signal_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def format_sheet(ws, freeze_panes='A2'):
        ws.freeze_panes = freeze_panes

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        format_sheet(ws)

        if sheet_name == 'README_Glossary' or sheet_name == 'Strategy_Overview':
            section_headers = ['PERFORMANCE METRICS:', 'RISK METRICS:', 'PAIR CHARACTERISTICS:',
                              'TRADING SIGNALS:', 'POSITION SIZING:', 'STRATEGY CONFIGURATION',
                              'PAIR SELECTION CRITERIA', 'STRATEGY LOGIC', 'RISK MANAGEMENT']

            for row in ws.iter_rows(min_row=2):
                if row[0].value in section_headers:
                    for cell in row:
                        cell.fill = title_fill
                        cell.font = title_font

        elif sheet_name == 'Current_Signals':
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if ws.cell(1, cell.column).value == 'Signal':
                        if cell.value in ['LONG SPREAD', 'SHORT SPREAD']:
                            for c in row:
                                c.fill = signal_fill

        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            header = col[0].value
            if header and isinstance(header, str):
                if '(%)' in header:
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        cell[0].number_format = '0.00'
                elif '(€)' in header or 'P&L' in header or 'Notional' in header:
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        cell[0].number_format = '€#,##0.00'
                elif 'Ratio' in header or 'β' in header or 'Correlation' in header or 'Factor' in header:
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        cell[0].number_format = '0.000'
                elif 'Trades' in header or 'Period' in header:
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        cell[0].number_format = '0'
                elif 'Z-Score' in header:
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        cell[0].number_format = '0.00'

    wb.save(output_file)

    print("\n Formatting complete!")
    print(f"\n Final enhanced file: {output_file}")
    print("\nSheets included:")
    print("  1. README_Glossary - Complete definitions with targets")
    print("  2. Portfolio_Pairs - Top performers with units and formatting")
    print("  3. All_Performance - All pairs ranked by Sharpe")
    print("  4. Current_Signals - Live trading signals (highlighted)")
    print("  5. Best_Pair_Detail - Daily backtest detail")
    print("  6. Correlation_Matrix - Top 50 bonds")
    print("  7. Portfolio_PnL - Aggregate returns")
    print("  8. Strategy_Overview - Full strategy documentation")
    print("  9. Portfolio_Summary - Quick stats dashboard")
    print("\nFile ready for trading desk presentation!")

except Exception as e:
    print(f"\n Error creating Excel file: {str(e)}")
    print("Creating another Excel file ")

    output_file = os.path.join(RESULTS_DIR, f'Pairs_Trading_Results_Enhanced_{nowstr}.xlsx')

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if len(portfolio_pairs) > 0:
                portfolio_pairs.to_excel(writer, sheet_name='Portfolio_Pairs', index=False)
            if len(performance_df) > 0:
                performance_df.to_excel(writer, sheet_name='All_Performance', index=False)
            if len(signals_df) > 0:
                signals_df.to_excel(writer, sheet_name='Current_Signals', index=False)
                
        excel_path = os.path.join(RESULTS_DIR, f'Pairs_Trading_Results_Enhanced_{nowstr}.xlsx')
        print(f"\n Basic Excel file created: {excel_path}")
    except Exception as ee:
        print(f" Could not create fallback Excel file: {ee}")


print("\nYou can analyze any specific pair by running:")
print("\n  analyze_custom_pair(isin1, isin2)")
print("\nExample:")
print("  analyze_custom_pair('ES0265936056 Corp', 'ES0265936023 Corp')")

def analyze_custom_pair(isin1, isin2, notional=1000000):

    print(f"\n{'='*80}")
    print(f"CUSTOM PAIR ANALYSIS: {isin1} vs {isin2}")
    print(f"{'='*80}")

    # Test cointegration
    coint_result = coint_tester.test_pair(isin1, isin2)

    print(f"\nCOINTEGRATION TEST:")
    print(f"  Beta (Hedge Ratio): {coint_result['Beta']:.3f}")
    print(f"  R-Squared: {coint_result['R_Squared']:.3f}")
    print(f"  Cointegration p-value: {coint_result['Coint_PValue']:.4f}")
    print(f"  Is Cointegrated: {coint_result['Is_Cointegrated']}")
    print(f"  Half-life: {coint_result['Half_Life']:.1f} days")
    print(f"  Current Z-Score: {coint_result['Current_ZScore']:.2f}")

    if coint_result['Is_Cointegrated']:
        pair_info = {
            'ISIN_1': isin1,
            'ISIN_2': isin2,
            'Alpha': coint_result['Alpha'],
            'Beta': coint_result['Beta']
        }
        bond1 = metadata[metadata['ISIN'] == isin1].iloc[0]
        bond2 = metadata[metadata['ISIN'] == isin2].iloc[0]
        pair_info['Duration_1'] = bond1['Mod Dur (Mid)']
        pair_info['Duration_2'] = bond2['Mod Dur (Mid)']

        bt_result = strategy.backtest_pair(pair_info, notional=notional)
        metrics = strategy.calculate_performance_metrics(bt_result)

        print(f"\nBACKTEST RESULTS:")
        print(f"  Sharpe Ratio: {metrics['Sharpe_Ratio']:.2f}")
        print(f"  Annualized Return: {metrics['Annualized_Return']*100:.2f}%")
        print(f"  Max Drawdown: {metrics['Max_Drawdown']*100:.2f}%")
        print(f"  Win Rate: {metrics['Win_Rate']*100:.1f}%")
        print(f"  Number of Trades: {int(metrics['Number_of_Trades'])}")

        fig, ax = plt.subplots(2, 1, figsize=(14, 8))

        plot_data = bt_result.dropna()

        ax[0].plot(plot_data['Date'], plot_data['ZScore'], linewidth=2)
        ax[0].axhline(2, color='g', linestyle='--', alpha=0.5)
        ax[0].axhline(-2, color='g', linestyle='--', alpha=0.5)
        ax[0].axhline(0.5, color='orange', linestyle='--', alpha=0.5)
        ax[0].axhline(-0.5, color='orange', linestyle='--', alpha=0.5)
        ax[0].set_title('Z-Score Evolution')
        ax[0].set_ylabel('Z-Score')
        ax[0].grid(alpha=0.3)

        ax[1].plot(plot_data['Date'], plot_data['PnL'], linewidth=2, color='darkgreen')
        ax[1].axhline(0, color='black', linestyle='-', linewidth=0.5)
        ax[1].set_title(f'Cumulative P&L: €{plot_data["PnL"].iloc[-1]:,.0f}')
        ax[1].set_ylabel('P&L (EUR)')
        ax[1].set_xlabel('Date')
        ax[1].grid(alpha=0.3)

        plt.tight_layout()
        plt.show()

    else:
        print("\n Pair is NOT cointegrated - not suitable for pairs trading")

    return coint_result

print("PAIRS TRADING SYSTEM")
print("="*80)

import time

REFRESH_INTERVAL_MINUTES = 5
ENABLE_LIVE_MONITORING = True  # Set to False to disable 

def refresh_data_and_signals():
    global signals_df, metadata
    return signals_df, metadata

while ENABLE_LIVE_MONITORING:
    print(f"\n\n Live Desk Monitor | {pd.Timestamp.now()}")
    try:
        signals_df, metadata = refresh_data_and_signals()
        print_trade_suggestions(signals_df, metadata)
        print_monitoring_pairs(signals_df)
        print_risk_summary(open_trades_df)
        print(f"Next update in {REFRESH_INTERVAL_MINUTES} minutes...")
        time.sleep(REFRESH_INTERVAL_MINUTES * 60)
    except KeyboardInterrupt:
        print("\nLive monitoring stopped by user.")
        break
    except Exception as e:
        print(f"Error during monitoring: {e}")
        time.sleep(30)

